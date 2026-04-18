from datetime import date, datetime, timedelta
from collections import defaultdict
import csv
from io import BytesIO
from io import StringIO

import bcrypt
import jwt
from fastapi import APIRouter, HTTPException, Request, status
from fastapi.responses import Response
from openpyxl import Workbook
from sqlalchemy import and_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import sessionmaker

from db import (
    AttendanceBase,
    AttendanceFillBase,
    AttendanceStatusEnum,
    ClassBase,
    RoleEnum,
    StudentBase,
    UserBase,
    engine,
)
from models import (
    AttendanceRequest,
    CreateClassRequest,
    LoginRequest,
    UpdateClassCredentialsRequest,
    UpdateCredentialsRequest,
    UpdateRoleRequest,
)
from utils.jwt import RANDOM_SECRET, create_jwt

router = APIRouter()
session = sessionmaker(engine)
HISTORY_RETENTION_DAYS = 7


def _normalize_absent_name(value: str) -> str:
    return " ".join(value.strip().split())


def _role_value(role: RoleEnum | str) -> str:
    return role.value if isinstance(role, RoleEnum) else str(role)


def _get_token_payload(request: Request) -> dict:
    auth_header = request.headers.get("Authorization")
    if not auth_header:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Unauthorized")
    parts = auth_header.split()
    if len(parts) != 2 or parts[0].lower() != "bearer":
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Unauthorized")
    try:
        return jwt.decode(parts[1], RANDOM_SECRET, algorithms=["HS256"])
    except jwt.PyJWTError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Unauthorized")


def _require_role(payload: dict, allowed: set[str]) -> None:
    role = payload.get("role")
    if role not in allowed:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")


def _get_current_user(s, payload: dict) -> UserBase:
    user = s.query(UserBase).filter(UserBase.id == int(payload["sub"])).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Unauthorized")
    return user


def _resolve_class_for_user(payload: dict, requested_class_id: int | None) -> int:
    if requested_class_id is not None:
        return requested_class_id
    if payload["role"] == RoleEnum.admin.value:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="classId is required for admin")
    with session() as s:
        owned_class = (
            s.query(ClassBase)
            .filter(ClassBase.teacher_id == int(payload["sub"]))
            .order_by(ClassBase.id.asc())
            .first()
        )
        if not owned_class:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Class not found")
        return owned_class.id


def _attendance_for_class(s, current_date: date, class_id: int) -> dict:
    attendance_rows = (
        s.query(AttendanceBase)
        .filter(and_(AttendanceBase.class_id == class_id, AttendanceBase.date == current_date))
        .all()
    )
    fill_row = (
        s.query(AttendanceFillBase)
        .filter(and_(AttendanceFillBase.class_id == class_id, AttendanceFillBase.date == current_date))
        .first()
    )
    is_filled = fill_row is not None
    total_students = fill_row.total_students if fill_row else 0
    present_count = fill_row.present_count if fill_row else 0
    unexcused = []
    excused = []
    for row in attendance_rows:
        if row.status == AttendanceStatusEnum.unexcused:
            unexcused.append({"fullName": row.absent_name})
        elif row.status == AttendanceStatusEnum.excused:
            excused.append(
                {
                    "fullName": row.absent_name,
                    "reason": row.reason or "",
                }
            )
    return {
        "date": current_date.isoformat(),
        "classId": class_id,
        "isFilled": is_filled,
        "totalStudents": total_students,
        "presentCount": present_count,
        "absentUnexcused": unexcused,
        "absentExcused": excused,
    }


def _daily_stats_for_class(class_row: ClassBase, absent_rows: list[AttendanceBase]) -> dict:
    class_id = class_row.id
    absent_list = []
    for row in absent_rows:
        reason = row.reason or "Неуважительная причина"
        absent_list.append(
            {
                "fullName": row.absent_name,
                "classId": class_id,
                "className": class_row.name,
                "reason": reason,
            }
        )
    return {
        "date": absent_rows[0].date.isoformat() if absent_rows else datetime.now().date().isoformat(),
        "classId": class_id,
        "className": class_row.name,
        "totalAbsent": len(absent_list),
        "absent": absent_list,
    }


def _attendance_for_classes(s, current_date: date, class_rows: list[ClassBase]) -> list[dict]:
    if not class_rows:
        return []
    class_ids = [row.id for row in class_rows]
    attendance_rows = (
        s.query(AttendanceBase)
        .filter(and_(AttendanceBase.date == current_date, AttendanceBase.class_id.in_(class_ids)))
        .all()
    )
    fill_rows = (
        s.query(AttendanceFillBase)
        .filter(and_(AttendanceFillBase.date == current_date, AttendanceFillBase.class_id.in_(class_ids)))
        .all()
    )
    fills_by_class = {row.class_id: row for row in fill_rows}
    attendance_by_class = defaultdict(list)
    for row in attendance_rows:
        attendance_by_class[row.class_id].append(row)

    result = []
    for class_row in class_rows:
        rows = attendance_by_class.get(class_row.id, [])
        fill_row = fills_by_class.get(class_row.id)
        is_filled = fill_row is not None
        total_students = fill_row.total_students if fill_row else 0
        present_count = fill_row.present_count if fill_row else 0
        unexcused = []
        excused = []
        for row in rows:
            if row.status == AttendanceStatusEnum.unexcused:
                unexcused.append({"fullName": row.absent_name})
            elif row.status == AttendanceStatusEnum.excused:
                excused.append({"fullName": row.absent_name, "reason": row.reason or ""})
        result.append(
            {
                "date": current_date.isoformat(),
                "classId": class_row.id,
                "isFilled": is_filled,
                "totalStudents": total_students,
                "presentCount": present_count,
                "absentUnexcused": unexcused,
                "absentExcused": excused,
            }
        )
    return result


def _cleanup_old_history(s) -> None:
    cutoff_date = datetime.now().date() - timedelta(days=HISTORY_RETENTION_DAYS - 1)
    s.query(AttendanceBase).filter(AttendanceBase.date < cutoff_date).delete()
    s.query(AttendanceFillBase).filter(AttendanceFillBase.date < cutoff_date).delete()


def _resolve_daily_stats_blocks(s, token_payload: dict, target_date: date, class_id: int | None) -> list[dict]:
    if class_id is None:
        if token_payload["role"] == RoleEnum.admin.value:
            class_rows = s.query(ClassBase).order_by(ClassBase.id.asc()).all()
        else:
            class_rows = (
                s.query(ClassBase)
                .filter(ClassBase.teacher_id == int(token_payload["sub"]))
                .order_by(ClassBase.id.asc())
                .all()
            )
        if not class_rows:
            return []
        class_ids = [row.id for row in class_rows]
        absent_rows = (
            s.query(AttendanceBase)
            .filter(
                and_(
                    AttendanceBase.date == target_date,
                    AttendanceBase.class_id.in_(class_ids),
                )
            )
            .all()
        )
        grouped_absent = defaultdict(list)
        for row in absent_rows:
            grouped_absent[row.class_id].append(row)
        blocks = []
        for row in class_rows:
            block = _daily_stats_for_class(row, grouped_absent.get(row.id, []))
            block["date"] = target_date.isoformat()
            blocks.append(block)
        return blocks

    resolved_class_id = _resolve_class_for_user(token_payload, class_id)
    class_row = s.query(ClassBase).filter(ClassBase.id == resolved_class_id).first()
    if not class_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Class not found")
    if token_payload["role"] == RoleEnum.teacher.value and class_row.teacher_id != int(token_payload["sub"]):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
    absent_rows = (
        s.query(AttendanceBase)
        .filter(
            and_(
                AttendanceBase.class_id == resolved_class_id,
                AttendanceBase.date == target_date,
            )
        )
        .all()
    )
    block = _daily_stats_for_class(class_row, absent_rows)
    block["date"] = target_date.isoformat()
    return [block]


@router.post("/auth/login")
def login(credentials: LoginRequest):
    with session() as s:
        user = s.query(UserBase).filter(UserBase.login == credentials.login).first()
        if not user:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")
        try:
            password_ok = bcrypt.checkpw(credentials.password.encode("utf-8"), user.password.encode("utf-8"))
        except ValueError:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")
        if not password_ok:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")
        role = _role_value(user.role)
        access_token = create_jwt({"sub": str(user.id), "role": role}, timedelta(days=7))
        return {"accessToken": access_token, "role": role, "userId": user.id}


@router.get("/users")
def get_users(request: Request):
    payload = _get_token_payload(request)
    _require_role(payload, {RoleEnum.admin.value})
    with session() as s:
        users = s.query(UserBase).order_by(UserBase.id.asc()).all()
        class_rows = s.query(ClassBase.teacher_id, ClassBase.id).all()
        class_map = {teacher_id: class_id for teacher_id, class_id in class_rows}
        return [
            {
                "id": user.id,
                "login": user.login,
                "role": _role_value(user.role),
                "classId": class_map.get(user.id),
                "promotedBy": user.promoted_by,
            }
            for user in users
        ]


@router.post("/users")
def register_teacher(request: Request):
    token_payload = _get_token_payload(request)
    _require_role(token_payload, {RoleEnum.admin.value})
    raise HTTPException(status_code=status.HTTP_410_GONE, detail="Teacher registration is disabled")


@router.patch("/users/{id}/credentials")
def update_credentials(id: int, request: Request, payload: UpdateCredentialsRequest):
    token_payload = _get_token_payload(request)
    _require_role(token_payload, {RoleEnum.admin.value})
    if payload.login is None and payload.password is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No update fields provided")
    with session() as s:
        teacher = s.query(UserBase).filter(and_(UserBase.id == id, UserBase.role == RoleEnum.teacher)).first()
        if not teacher:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Teacher not found")
        if payload.login is not None:
            teacher.login = payload.login
        if payload.password is not None:
            teacher.password = bcrypt.hashpw(payload.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Login already exists")
        return {"message": "Updated"}


@router.patch("/profile/credentials")
def update_my_credentials(request: Request, payload: UpdateCredentialsRequest):
    token_payload = _get_token_payload(request)
    if payload.login is None and payload.password is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No update fields provided")
    with session() as s:
        current_user = _get_current_user(s, token_payload)
        if payload.login is not None:
            current_user.login = payload.login
        if payload.password is not None:
            current_user.password = bcrypt.hashpw(payload.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Login already exists")
        return {"message": "Updated"}


@router.patch("/users/{id}/role")
def update_user_role(id: int, request: Request, payload: UpdateRoleRequest):
    token_payload = _get_token_payload(request)
    _require_role(token_payload, {RoleEnum.admin.value})
    if payload.role not in {RoleEnum.teacher.value, RoleEnum.admin.value}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid role")
    with session() as s:
        actor = _get_current_user(s, token_payload)
        target = s.query(UserBase).filter(UserBase.id == id).first()
        if not target:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

        if actor.promoted_by == target.id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")

        new_role = RoleEnum(payload.role)
        target.role = new_role
        if new_role == RoleEnum.admin:
            target.promoted_by = actor.id
        else:
            target.promoted_by = None
        s.commit()
        return {"message": "Updated"}


@router.get("/classes")
def get_classes(request: Request):
    payload = _get_token_payload(request)
    with session() as s:
        if payload["role"] == RoleEnum.admin.value:
            class_rows = s.query(ClassBase).all()
        else:
            class_rows = s.query(ClassBase).filter(ClassBase.teacher_id == int(payload["sub"])).all()
        return [{"id": row.id, "name": row.name, "teacherId": row.teacher_id} for row in class_rows]


@router.post("/classes", status_code=status.HTTP_201_CREATED)
def create_class(request: Request, payload: CreateClassRequest):
    token_payload = _get_token_payload(request)
    _require_role(token_payload, {RoleEnum.admin.value})
    with session() as s:
        class_login = payload.name.strip()
        if not class_login:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Class name is required")
        if not payload.password:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Password is required")
        hashed_password = bcrypt.hashpw(payload.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        class_user = UserBase(login=class_login, password=hashed_password, role=RoleEnum.teacher)
        class_row = ClassBase(name=class_login, teacher_id=None)
        s.add(class_user)
        s.flush()
        class_row.teacher_id = class_user.id
        s.add(class_row)
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Class name or login already exists")
        return {"message": "Class created"}


@router.patch("/classes/{id}/credentials")
def update_class_credentials(id: int, request: Request, payload: UpdateClassCredentialsRequest):
    token_payload = _get_token_payload(request)
    _require_role(token_payload, {RoleEnum.admin.value})
    if payload.login is None and payload.password is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No update fields provided")
    with session() as s:
        class_row = s.query(ClassBase).filter(ClassBase.id == id).first()
        if not class_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Class not found")
        class_user = (
            s.query(UserBase)
            .filter(and_(UserBase.id == class_row.teacher_id, UserBase.role == RoleEnum.teacher))
            .first()
        )
        if not class_user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Class account not found")
        if payload.login is not None:
            new_login = payload.login.strip()
            if not new_login:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Login cannot be empty")
            class_user.login = new_login
            class_row.name = new_login
        if payload.password is not None:
            if not payload.password:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Password cannot be empty")
            class_user.password = bcrypt.hashpw(payload.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Class name or login already exists")
        return {"message": "Updated"}


@router.delete("/classes/{id}")
def delete_class(id: int, request: Request):
    token_payload = _get_token_payload(request)
    _require_role(token_payload, {RoleEnum.admin.value})
    with session() as s:
        class_row = s.query(ClassBase).filter(ClassBase.id == id).first()
        if not class_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Class not found")
        class_user_id = class_row.teacher_id
        s.query(AttendanceBase).filter(AttendanceBase.class_id == id).delete()
        s.query(AttendanceFillBase).filter(AttendanceFillBase.class_id == id).delete()
        s.query(StudentBase).filter(StudentBase.class_id == id).delete()
        s.delete(class_row)
        if class_user_id is not None:
            class_user = (
                s.query(UserBase)
                .filter(and_(UserBase.id == class_user_id, UserBase.role == RoleEnum.teacher))
                .first()
            )
            if class_user:
                s.delete(class_user)
        s.commit()
        return {"message": "Deleted"}


@router.get("/attendance")
def get_attendance(date: date, request: Request, classId: int | None = None):
    token_payload = _get_token_payload(request)
    with session() as s:
        _cleanup_old_history(s)
        s.commit()
        if classId is None and token_payload["role"] == RoleEnum.admin.value:
            class_rows = s.query(ClassBase).order_by(ClassBase.id.asc()).all()
            return _attendance_for_classes(s, date, class_rows)

        resolved_class_id = _resolve_class_for_user(token_payload, classId)
        class_row = s.query(ClassBase).filter(ClassBase.id == resolved_class_id).first()
        if not class_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Class not found")
        if token_payload["role"] == RoleEnum.teacher.value and class_row.teacher_id != int(token_payload["sub"]):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
        return _attendance_for_class(s, date, resolved_class_id)


@router.put("/attendance")
def put_attendance(date: date, request: Request, payload: AttendanceRequest):
    token_payload = _get_token_payload(request)
    with session() as s:
        _cleanup_old_history(s)
        s.commit()
        resolved_class_id = _resolve_class_for_user(token_payload, payload.class_id)
        class_row = s.query(ClassBase).filter(ClassBase.id == resolved_class_id).first()
        if not class_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Class not found")
        if token_payload["role"] == RoleEnum.teacher.value and class_row.teacher_id != int(token_payload["sub"]):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
        if payload.total_students < 0 or payload.present_count < 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid totals")
        if payload.present_count > payload.total_students:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Present count exceeds total")
        absent_unexcused = []
        for name in payload.absent_unexcused:
            normalized_name = _normalize_absent_name(name)
            if normalized_name:
                absent_unexcused.append(normalized_name)

        absent_excused = []
        for item in payload.absent_excused:
            normalized_name = _normalize_absent_name(item.full_name)
            if not normalized_name:
                continue
            absent_excused.append({"fullName": normalized_name, "reason": item.reason.strip()})
        if any(not item["reason"] for item in absent_excused):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Reason is required for excused absence")

        normalized_unexcused = {name.casefold() for name in absent_unexcused}
        normalized_excused = {item["fullName"].casefold() for item in absent_excused}
        if len(normalized_unexcused) != len(absent_unexcused) or len(normalized_excused) != len(absent_excused):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Duplicate absent fullName")
        if normalized_unexcused & normalized_excused:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Duplicate absent fullName")

        absent_count = len(absent_unexcused) + len(absent_excused)
        expected_absent = payload.total_students - payload.present_count
        if absent_count != expected_absent:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Absent count must match totalStudents - presentCount",
            )

        s.query(AttendanceBase).filter(
            and_(AttendanceBase.date == date, AttendanceBase.class_id == resolved_class_id)
        ).delete()

        for name in absent_unexcused:
            s.add(
                AttendanceBase(
                    date=date,
                    class_id=resolved_class_id,
                    absent_name=name,
                    status=AttendanceStatusEnum.unexcused,
                )
            )
        for item in absent_excused:
            s.add(
                AttendanceBase(
                    date=date,
                    class_id=resolved_class_id,
                    absent_name=item["fullName"],
                    status=AttendanceStatusEnum.excused,
                    reason=item["reason"],
                )
            )

        existing_fill = (
            s.query(AttendanceFillBase)
            .filter(and_(AttendanceFillBase.date == date, AttendanceFillBase.class_id == resolved_class_id))
            .first()
        )
        if not existing_fill:
            s.add(
                AttendanceFillBase(
                    date=date,
                    class_id=resolved_class_id,
                    total_students=payload.total_students,
                    present_count=payload.present_count,
                )
            )
        else:
            existing_fill.total_students = payload.total_students
            existing_fill.present_count = payload.present_count
            existing_fill.filled_at = datetime.now()
        s.commit()
        return {"message": "Saved"}


@router.get("/statistics/daily")
def get_daily_statistics(date: date, request: Request, classId: int | None = None):
    token_payload = _get_token_payload(request)
    with session() as s:
        _cleanup_old_history(s)
        s.commit()
        blocks = _resolve_daily_stats_blocks(s, token_payload, date, classId)
        if classId is None:
            return blocks
        return blocks[0]


@router.get("/statistics/daily/export")
def export_daily_statistics_excel(date: date, request: Request, classId: int | None = None):
    token_payload = _get_token_payload(request)
    with session() as s:
        _cleanup_old_history(s)
        s.commit()
        blocks = _resolve_daily_stats_blocks(s, token_payload, date, classId)

    workbook = Workbook()
    details_sheet = workbook.active
    details_sheet.title = "Absent details"
    details_sheet.append(["Date", "Class ID", "Class Name", "Full Name", "Reason"])

    for block in blocks:
        for item in block["absent"]:
            details_sheet.append(
                [
                    block["date"],
                    block["classId"],
                    block["className"],
                    item["fullName"],
                    item["reason"],
                ]
            )

    if details_sheet.max_row == 1:
        details_sheet.append([date.isoformat(), "-", "-", "No absences", "-"])

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Date", "Class ID", "Class Name", "Total absent"])
    total_absent_all = 0
    for block in blocks:
        summary_sheet.append([block["date"], block["classId"], block["className"], block["totalAbsent"]])
        total_absent_all += block["totalAbsent"]
    summary_sheet.append([date.isoformat(), "-", "TOTAL", total_absent_all])

    details_sheet.column_dimensions["A"].width = 14
    details_sheet.column_dimensions["B"].width = 10
    details_sheet.column_dimensions["C"].width = 24
    details_sheet.column_dimensions["D"].width = 28
    details_sheet.column_dimensions["E"].width = 28
    summary_sheet.column_dimensions["A"].width = 14
    summary_sheet.column_dimensions["B"].width = 10
    summary_sheet.column_dimensions["C"].width = 24
    summary_sheet.column_dimensions["D"].width = 14

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    class_suffix = f"_class_{classId}" if classId is not None else "_all_classes"
    filename = f"attendance_statistics_{date.isoformat()}{class_suffix}.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/statistics/daily/export/csv")
def export_daily_statistics_csv(date: date, request: Request, classId: int | None = None):
    token_payload = _get_token_payload(request)
    with session() as s:
        _cleanup_old_history(s)
        s.commit()
        blocks = _resolve_daily_stats_blocks(s, token_payload, date, classId)

    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(["Date", "Class ID", "Class Name", "Full Name", "Reason"])

    has_absences = False
    for block in blocks:
        for item in block["absent"]:
            writer.writerow([block["date"], block["classId"], block["className"], item["fullName"], item["reason"]])
            has_absences = True

    if not has_absences:
        writer.writerow([date.isoformat(), "-", "-", "No absences", "-"])

    writer.writerow([])
    writer.writerow(["Summary"])
    writer.writerow(["Date", "Class ID", "Class Name", "Total absent"])
    total_absent_all = 0
    for block in blocks:
        writer.writerow([block["date"], block["classId"], block["className"], block["totalAbsent"]])
        total_absent_all += block["totalAbsent"]
    writer.writerow([date.isoformat(), "-", "TOTAL", total_absent_all])

    class_suffix = f"_class_{classId}" if classId is not None else "_all_classes"
    filename = f"attendance_statistics_{date.isoformat()}{class_suffix}.csv"

    return Response(
        content=csv_buffer.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/attendance/unfilled-classes")
def get_unfilled_classes(date: date, request: Request):
    token_payload = _get_token_payload(request)
    with session() as s:
        _cleanup_old_history(s)
        s.commit()
        if token_payload["role"] == RoleEnum.admin.value:
            class_rows = s.query(ClassBase).order_by(ClassBase.id.asc()).all()
        else:
            class_rows = (
                s.query(ClassBase)
                .filter(ClassBase.teacher_id == int(token_payload["sub"]))
                .order_by(ClassBase.id.asc())
                .all()
            )
        filled_class_ids = {
            class_id
            for (class_id,) in s.query(AttendanceFillBase.class_id).filter(AttendanceFillBase.date == date).all()
        }
        teacher_ids = {row.teacher_id for row in class_rows}
        teacher_rows = s.query(UserBase.id, UserBase.login).filter(UserBase.id.in_(teacher_ids)).all() if teacher_ids else []
        teacher_map = {teacher_id: login for teacher_id, login in teacher_rows}
        return [
            {
                "id": row.id,
                "name": row.name,
                "teacherId": row.teacher_id,
                "teacherLogin": teacher_map.get(row.teacher_id),
            }
            for row in class_rows
            if row.id not in filled_class_ids
        ]
